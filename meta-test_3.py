

import os
from random import randint
import re
import threading
from time import sleep
import requests
from bs4 import BeautifulSoup
import xlsxwriter
from openpyxl.drawing.image import Image

from openpyxl import load_workbook
from openpyxl import Workbook
# Load in the workbook

def load_xls (name):
    try:
        wb = load_workbook(f'./{name}.xlsx')
    except:
        wb = Workbook()
        sheet = wb.active 
        sheet['A1'] = 'Название игры'
        sheet['B1'] = 'Платформа'
        sheet['C1'] = 'Пользовательская оценка'
        sheet['D1'] = 'Оценка критиков'
        wb.save(f'./{name}.xlsx')

    wb = load_workbook(f'./{name}.xlsx')


    sheet = wb[(wb.sheetnames[0])]


    row=[]
    str=sheet['A1'].value
    for i in range (2, 100000):
        #print(str)
        row.append(str)
        str = sheet[f'A{i}'].value
        if str==None: break
    return row




ps_game_ls=[]
games = []

class game:
    def __init__(self, name, link, platform, critic_score, user_score = None, poster_link = None):
        self.name = name
        self.link = link
        self.platform = platform
        self.critic_score = critic_score
        self.user_score = user_score
        self.poster_link = poster_link

    def __repr__(self):
        return f'\nName: {self.name}\nLink: {self.link}\nPlatform: {self.platform}\nUser score: {self.user_score}\nCritic score: {self.critic_score}\nPoster link: {self.poster_link}\n'



def get_ps_game():
    url = "https://www.playstation.com/ru-ua/ps-plus/games/"
    r = requests.get(url)
    global ps_game_ls
    
    alp='A B C D E F G H I J K L M N O P Q R S T U V W X Y Z 0 '

    soup = BeautifulSoup (r.text, 'html.parser')
    for i in soup.find_all("div", {"class": "txt-block-paragraph text-align--left"}):
        z=i.get_text().strip()

        if z[0:1] in alp:
            game = z[3:].strip().replace('*','')
            print (game)
            games = game.split('\n')
            ps_game_ls+=games

def go(str):
    url = "https://www.metacritic.com"
    loc = '/game/playstation-5/nba-2k21'
    f_url = url+loc
    seen = set()
    while True:
        r = requests.get(f_url, allow_redirects=False)
        loc = r.headers['location']
        if 'metacritic' in loc:
            f_url = loc
        else: 
            f_url = url+loc
        if loc in seen: break
        seen.add(loc)
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.131 Safari/537.36'
    r = s.get(f_url)
    return r.text

#print (go('a'))


def search_game(str:str):
    url = "https://www.metacritic.com"
    str = str.strip().replace('- ','').replace('’','').replace(':','').replace(' ','-').replace('\xa0','').lower()
    str = re.sub(r'\s\D* Edition', '', str)
    str = re.sub(r'\s\D* Version', '', str)
    loc = f'/search/game/{str}/results'
    f_url = url+loc
    seen = set()
    p=0
    while p<10000:
        r = requests.get(f_url, allow_redirects=False)
        loc = r.headers['location']
        if 'metacritic' in loc:
            f_url = loc
        else: 
            f_url = url+loc
        if loc in seen: break
        seen.add(loc)
        p+=1
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.131 Safari/537.36'
    r = s.get(f_url)
    #sleep(2)
    print(r.status_code)
    soup = BeautifulSoup (r.text, 'html.parser')
    return soup


def get_user_score (link):
    url = "https://www.metacritic.com"
    loc = link.replace(url,'')
    f_url = url+loc
    seen = set()
    while True:
        r = requests.get(f_url, allow_redirects=False)
        loc = r.headers['location']
        if 'metacritic' in loc:
            f_url = loc
        else: 
            f_url = url+loc
        if loc in seen: break
        seen.add(loc)
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.131 Safari/537.36'
    r = s.get(f_url)
    #sleep(2)
    soup = BeautifulSoup (r.text, 'html.parser')
    user_score = soup.find("div", {"class": re.compile("^metascore_w user large game")}).text.replace('.','')
    poster_link = soup.find("img", {"class": re.compile("^product_image")}).get('src')
    return user_score, poster_link

def claster (soup:BeautifulSoup):
    base_url = "https://www.metacritic.com"
    for s2 in soup.find_all("div", {"class": "main_stats"}):
        try:
            mark = s2.find("span", {"class": re.compile("^metascore_w medium game")}).text
            platform = s2.find("span", {"class": "platform"}).text
            name = s2.find('a')
            link = base_url + name.get('href')
            user_score_link = get_user_score(link)
            name = name.text.strip()
            #games.append(game(name, link, platform, mark))
            if platform in ['PS4', 'PS5']:
                return [link, platform, mark, *user_score_link]
        except:
            pass
    return [None, None, None, None, None]


def _excel(with_img):   
    # открываем новый файл на запись
    workbook = xlsxwriter.Workbook(f'Ps_plus{"_with_img" if with_img else ""}.xlsx')
    # создаем там "лист"
    worksheet = workbook.add_worksheet()
    # в ячейку A1 пишем текст
    worksheet.write('A1', 'Название игры')
    worksheet.write('B1', 'Платформа')
    worksheet.write('C1', 'Пользовательская оценка')
    worksheet.write('D1', 'Оценка критиков')
    if with_img: 
        worksheet.write('E1', 'Постер')
    
    for i in range (1, 1 + len(games)):
        try:
            worksheet.write_url(i, 0, games[i-1].link, string=games[i-1].name)
            worksheet.write(i, 1, games[i-1].platform)
            worksheet.write(i, 2, int(games[i-1].user_score))
            worksheet.write(i, 3, int(games[i-1].critic_score))

            if with_img: 
                a=requests.get(games[i-1].poster_link)
                with open(f'{i}.jpg', 'wb') as file:
                    file.write(a.content)
                worksheet.insert_image(i, 4, f'{i}.jpg')
        except:
            pass

        
        
    workbook.close()

def excel(name, with_img):
    row_len = 0   
    wb = load_workbook(f'./{name}.xlsx')

    sheet = wb.active


    row=[]
    str=sheet['A1'].value
    for i in range (2, 100000):
        #print(str)
        row.append(str)
        str = sheet[f'A{i}'].value
        row_len = i
        if str==None: break

    
    
    for i in range (row_len, row_len + len(games)):
        
        sheet[f'A{i}'] = games[row_len-i].name
        sheet[f'B{i}'] = games[row_len-i].link
        sheet[f'C{i}'] = games[row_len-i].platform
        try:
            sheet[f'D{i}'] = int(games[row_len-i].user_score)
        except: 
            pass
        
        try:
            sheet[f'E{i}'] = int(games[row_len-i].critic_score)
        except: 
            pass

        if with_img: 
            a=requests.get(games[row_len-i].poster_link)
            with open(f'{i}.jpg', 'wb') as file:
                file.write(a.content)
            img = Image(f'{i}.jpg')
            wb.add_image(img, f'F{i}')

    wb.save(filename = f'./{name}.xlsx')

            #sheet[] = games[i-1].name
            #worksheet.insert_image(i, 4, f'{i}.jpg')
    

        
        


def start ():
    global ps_game_ls
    get_ps_game()
    # ps_game_ls=['Agatha Christie - The ABC Murders', 'Danger Zone', 'Dandara: Trials of Fear Edition', 'Embr', 'Far Cry 3 Blood Dragon: Classic Edition', 'FIA European Truck Racing Championship', 'FIGHTING EX LAYER - Standard Version', 'Ice Age: Scrat’s Nutty Adventure', 'Left Alive: Day One Edition', 'Marvel’s Avengers', 'Observation', 'Sine Mora X', 'The Artful Escape', 'The Council - The Complete Season', 'Warhammer: Chaosbane - Slayer Edition\xa0']
    # ps_game_ls=ps_game_ls[0:2]
    row = load_xls('Ps_plus')
    bad_game=[]

    
    tmp = False
    for i in range(len(ps_game_ls)):
        if ps_game_ls[i][0] != 'A':
            tmp = True
        if tmp and ps_game_ls[i][0] == 'A' and  ps_game_ls[i+1][0] == 'A':
            ps_game_ls = ps_game_ls[:i]
            break
    print (ps_game_ls)


    for ps_game in ps_game_ls:
        
        if ps_game in row:
            print (f'Skip {ps_game}')
            continue

        p=0
        while True:
            g=claster(search_game(ps_game))
            if g[0]==None:
                print (g)
                print (p)
                sleep(3)
                p+=1
                if p > 3:
                    bad_game.append(ps_game)
                    break
                continue
            else: 
                break
        if g[0]!=None:
            games.append(game(ps_game, *g))
            print (games[-1])
    print (games)

    #excel(True)
    excel('Ps_plus', False)
    #excel('Ps_plus_with_img', True)

    print ('bad game')
    print (bad_game)

start()


#print (claster(search_game('Adr1ft')))